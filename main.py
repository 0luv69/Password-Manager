import time
import random
import string
import os
import openpyxl
import xlsxwriter

headder = ["S.n","Date","Name","Password"]
CurrentLocation=(os.path.dirname(__file__)).replace("\\","/")
def reader():
    values=[]

    try:    
        book = openpyxl.load_workbook(f"{CurrentLocation}/save_folder.xlsx")
    except :
        book = openpyxl.Workbook()
        book.save(f"{CurrentLocation}/save_folder.xlsx")
        return values

    sheet = book.active
    rows = sheet.rows
    values=[]
    while True:
        try:
            values.append([cell.value for cell in next(rows) ])
        except:
            break 
    values.pop(0)     
    return values  



def writters(data):
    books = xlsxwriter.Workbook(f"{CurrentLocation}/save_folder.xlsx")
    sheet1 = books.add_worksheet('SHEET1')
    a=1
    for index1,hedder in enumerate (headder):
        sheet1.write(0,index1,hedder)
    for index2,entry1 in enumerate(data):
        for index3,entry2 in enumerate(entry1):
            sheet1.write(a,index3,entry2)
        a+=1
    books.close()    

if __name__ == "__main__":

    while True:
        ask = int(input("Enter the Length of your password:\t"))
        if ask >=8:
            break
        else:
            print("\nSorry plz enter more then 8 long degit code")

    password8 = str(random.randint(0,9))+("".join(random.choice(string.ascii_letters) for i in range(int(2)))) + str(random.randint(0,9)) + ("".join(random.choice(string.ascii_letters) for i in range(int(3)))) +str(random.randint(0,9))
    newpassword = password8 +("".join(random.choice(string.ascii_letters) for i in range(ask-8)))
    print(f"Your password is:\t\t{newpassword}\n")
    save= input("Do you want to save the password:[y/n]\t")
    if save.upper() == "Y":
        name = (input("Enter the name:\t\t")).capitalize()
        times = (time.strftime("%m/%d/%Y, %H:%M:%S"))
        first_d = reader()
        s = len(first_d)
        second_d =first_d+ [[s,times,name,newpassword]]
        writters(second_d)
    else:
        print("Thanks for chosing us")    




    


